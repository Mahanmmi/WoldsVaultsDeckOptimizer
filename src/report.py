"""Reporting & output: heatmaps, HNS metric, terminal report, xlsx export."""
from __future__ import annotations

import math
from collections import Counter
from typing import Dict, FrozenSet, List, Tuple

from .types import (
    CardClass, CardType, CoreType, Position,
    GREED_TYPES, REGULAR_TYPES, DELUXE_TYPES, TYPELESS_TYPES,
    PLACEABLE,
)
from .config import (
    ADDITIVE_CORES,
    ALLOW_DELUXE,
    BALANCE_DISPLAY,
    Deck,
    ENABLE_EXPERIMENTAL_EXPONENT,
    EXPERIMENTAL_BOOST,
    EXPERIMENTAL_EXPONENT,
    GREED_ADDITIVE,
    HEATMAP_DISPLAY,
    HNS_Q,
    MULT_COLOR,
    MULT_DELUXE_CORE_BASE,
    MULT_DELUXE_CORE_SCALE,
    MULT_DELUXE_FLAT,
    MULT_DIR_GREED_DIAG_DOWN,
    MULT_DIR_GREED_DIAG_UP,
    MULT_DIR_GREED_HORIZ,
    MULT_DIR_GREED_VERT,
    MULT_EQUILIBRIUM,
    MULT_EVO_GREED,
    MULT_FOIL,
    MULT_PURE_BASE,
    MULT_PURE_SCALE,
    MULT_STEADFAST,
    MULT_SURR_GREED,
    SHINY_POSITIONAL,
    SPREAD_METRIC,
    _get_test_configs,
)
from .simulate import _apply_greed

# ──────────────────────────────────────────────────────────────────────────────
# Post-processing helpers
# ──────────────────────────────────────────────────────────────────────────────

def mark_filler_greed(
    deck: Deck, assignment: Dict[Position, CardType], card_class: CardClass,
) -> Dict[Position, CardType]:
    regular = frozenset(p for p, t in assignment.items()
                        if t in REGULAR_TYPES | DELUXE_TYPES | TYPELESS_TYPES)
    result  = dict(assignment)
    for p, t in assignment.items():
        if t not in GREED_TYPES or t == CardType.FILLER_GREED:
            continue
        gr, gc  = p
        buffing = False
        if   t == CardType.DIR_GREED_UP:    buffing = (gr-1, gc) in regular
        elif t == CardType.DIR_GREED_DOWN:   buffing = (gr+1, gc) in regular
        elif t == CardType.DIR_GREED_LEFT:   buffing = (gr, gc-1) in regular
        elif t == CardType.DIR_GREED_RIGHT:  buffing = (gr, gc+1) in regular
        elif t == CardType.DIR_GREED_NE:  buffing = (gr - 1, gc + 1) in regular
        elif t == CardType.DIR_GREED_NW:  buffing = (gr - 1, gc - 1) in regular
        elif t == CardType.DIR_GREED_SE:  buffing = (gr + 1, gc + 1) in regular
        elif t == CardType.DIR_GREED_SW:  buffing = (gr + 1, gc - 1) in regular
        elif t == CardType.EVO_GREED:        buffing = card_class == CardClass.EVO and (gr+1, gc) in regular
        elif t == CardType.SURR_GREED:       buffing = any(q in regular for q in deck._surr_peers[p])
        if not buffing:
            result[p] = CardType.FILLER_GREED
    return result


def compute_heatmap(
    deck:       Deck,
    assignment: Dict[Position, CardType],
    card_class: CardClass,
    cores:      FrozenSet[CoreType],
) -> Dict[Position, float]:
    greed:    Dict[Position, CardType] = {}
    regular:  Dict[Position, CardType] = {}
    deluxe:   Dict[Position, CardType] = {}
    typeless: Dict[Position, CardType] = {}

    for p, t in assignment.items():
        if   t in GREED_TYPES:    greed[p]    = t
        elif t in REGULAR_TYPES:  regular[p]  = t
        elif t in DELUXE_TYPES:   deluxe[p]   = t
        elif t in TYPELESS_TYPES: typeless[p] = t

    filled   = frozenset(greed) | frozenset(regular) | frozenset(deluxe) | frozenset(typeless)
    scorable = {**regular, **deluxe, **typeless}

    foil_active = CoreType.FOIL in cores
    if card_class == CardClass.EVO:
        n_ns = len(greed) if foil_active else (len(regular) + len(greed))
    else:
        n_ns = len(greed)
    n_deluxe = len(deluxe)

    core_contributions        = []
    deluxe_core_contributions = []
    for core in cores:
        if   core == CoreType.PURE:
            core_contributions.append(MULT_PURE_BASE + MULT_PURE_SCALE * (n_ns + deck.n_arcane))
        elif core == CoreType.EQUILIBRIUM and card_class == CardClass.SHINY:
            core_contributions.append(MULT_EQUILIBRIUM)
        elif core == CoreType.STEADFAST   and card_class == CardClass.SHINY:
            core_contributions.append(MULT_STEADFAST)
        elif core == CoreType.COLOR:
            core_contributions.append(MULT_COLOR)
        elif core == CoreType.FOIL:
            core_contributions.append(MULT_FOIL)
        elif core == CoreType.DELUXE_CORE:
            deluxe_core_contributions.append(
                MULT_DELUXE_CORE_BASE + MULT_DELUXE_CORE_SCALE * n_deluxe)

    if ADDITIVE_CORES:
        core_mult        = 1.0 + sum(v - 1.0 for v in core_contributions)        if core_contributions        else 1.0
        deluxe_core_mult = 1.0 + sum(v - 1.0 for v in deluxe_core_contributions) if deluxe_core_contributions else 1.0
    else:
        core_mult        = math.prod(core_contributions)        if core_contributions        else 1.0
        deluxe_core_mult = math.prod(deluxe_core_contributions) if deluxe_core_contributions else 1.0

    row_count: Dict[int, int] = {}
    col_count: Dict[int, int] = {}
    for r, c in filled:
        row_count[r] = row_count.get(r, 0) + 1
        col_count[c] = col_count.get(c, 0) + 1

    init  = 1.0
    boost = {p: init for p in scorable}

    for g, gt in greed.items():
        gr, gc = g
        if   gt == CardType.DIR_GREED_UP:
            t = (gr - 1, gc)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_VERT)
        elif gt == CardType.DIR_GREED_DOWN:
            t = (gr + 1, gc)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_VERT)
        elif gt == CardType.DIR_GREED_LEFT:
            t = (gr, gc - 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_HORIZ)
        elif gt == CardType.DIR_GREED_RIGHT:
            t = (gr, gc + 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_HORIZ)
        elif gt == CardType.DIR_GREED_NE:
            t = (gr - 1, gc + 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_DIAG_UP)
        elif gt == CardType.DIR_GREED_NW:
            t = (gr - 1, gc - 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_DIAG_UP)
        elif gt == CardType.DIR_GREED_SE:
            t = (gr + 1, gc + 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_DIAG_DOWN)
        elif gt == CardType.DIR_GREED_SW:
            t = (gr + 1, gc - 1)
            if t in scorable:  _apply_greed(boost, t, MULT_DIR_GREED_DIAG_DOWN)
        elif gt == CardType.EVO_GREED:
            if card_class == CardClass.EVO:
                t = (gr + 1, gc)
                if t in regular:
                    _apply_greed(boost, t, MULT_EVO_GREED)
        elif gt == CardType.SURR_GREED:
            for t in deck._surr_peers[g]:
                if t in scorable:  _apply_greed(boost, t, MULT_SURR_GREED)

    heatmap: Dict[Position, float] = {p: 0.0 for p in deck.slots}

    def _contrib(val: float) -> float:
        return (val * EXPERIMENTAL_BOOST) ** EXPERIMENTAL_EXPONENT if ENABLE_EXPERIMENTAL_EXPONENT else val

    for p, t in regular.items():
        r, c = p
        if   t == CardType.ROW:  pos = row_count.get(r, 0)
        elif t == CardType.COL:  pos = col_count.get(c, 0)
        elif t == CardType.DIAG: pos = sum(1 for q in deck._diag_peers[p]
                                           if q in filled) + 1
        else:                    pos = sum(1 for q in deck._surr_peers[p] if q in filled)
        b          = max(boost[p], 1.0) if GREED_ADDITIVE else boost[p]
        heatmap[p] = _contrib(pos * core_mult * deluxe_core_mult * b)

    for p in deluxe:
        b          = max(boost[p], 1.0) if GREED_ADDITIVE else boost[p]
        heatmap[p] = _contrib(MULT_DELUXE_FLAT * core_mult * b)

    for p in typeless:
        b          = max(boost[p], 1.0) if GREED_ADDITIVE else boost[p]
        heatmap[p] = _contrib(1.0 * core_mult * deluxe_core_mult * b)

    return heatmap


def compute_hns(heatmap: Dict[Position, float], q: float) -> float:
    """
    Hill Number Score = N · (Σ pᵢ^q)^(1/(1-q))
    where pᵢ = hᵢ / N are the fractional NDM contributions of regular cards.
    At q=1 (limit): N · exp(-Σ pᵢ ln pᵢ)  (Shannon-weighted NDM).
    At q=2:         N² / Σ hᵢ²             (Effective Deck Power / Simpson reciprocal).
    Only positive heatmap values are included (greed/empty slots excluded).
    """
    vals = [v for v in heatmap.values() if v > 0]
    if not vals:
        return 0.0
    N = sum(vals)
    if N == 0:
        return 0.0
    ps = [v / N for v in vals]
    if abs(q - 1.0) < 1e-9:   # limit: Shannon
        return N * math.exp(-sum(p * math.log(p) for p in ps))
    return N * (sum(p ** q for p in ps)) ** (1.0 / (1.0 - q))


def compute_hns_from_result(deck: Deck, result: dict, card_class: CardClass) -> float:
    heatmap = compute_heatmap(deck, result["assignment"], card_class, result["cores"])
    return compute_hns(heatmap, HNS_Q)


# ──────────────────────────────────────────────────────────────────────────────
# Terminal report
# ──────────────────────────────────────────────────────────────────────────────

def _report(card_class: CardClass, best: dict, deck: Deck) -> None:
    if best["score"] < 0:
        print("  (no valid configuration found)"); return

    asgn  = best["assignment"]
    cores = best["cores"]

    display_asgn = mark_filler_greed(deck, asgn, card_class)
    counts       = Counter(display_asgn.values())
    n_reg        = sum(counts[t] for t in REGULAR_TYPES)
    n_grd        = sum(counts[t] for t in GREED_TYPES)

    dashes = '─' * max(0, 44 - len(deck.name))
    print(f"\n  ── Best for {card_class.value.upper()}: {deck.name} {dashes}")
    print(f"  NDM           : {best['score']:.4f}")
    print(f"  Cores         : {sorted(c.value for c in cores)}")
    print(f"  Regular cards : {n_reg}   Greed cards : {n_grd}")
    print(f"  Constraint    : {deck.constraint_str()}")
    breakdown = "  ".join(
        f"{t.value}={counts[t]}" for t in PLACEABLE + [CardType.FILLER_GREED] if counts[t] > 0)
    print(f"  Breakdown     : {breakdown}")

    if BALANCE_DISPLAY:
        print(f"\n  ── Balance info {'─'*51}")
        print(f"  Greed stacking  : {'additive' if GREED_ADDITIVE else 'multiplicative'}")
        print(f"  surr_greed      : {MULT_SURR_GREED}x")
        print(f"  dir_greed_vert  : {MULT_DIR_GREED_VERT}x")
        print(f"  dir_greed_horiz : {MULT_DIR_GREED_HORIZ}x")
        print(f"  evo_greed       : {MULT_EVO_GREED}x")
        print(f"  pure_core       : {MULT_PURE_BASE} + {MULT_PURE_SCALE} × n_ns")
        print(f"  equilibrium     : {MULT_EQUILIBRIUM}x  (shiny only)")
        print(f"  steadfast       : {MULT_STEADFAST}x  (shiny only)")
        print(f"  color           : {MULT_COLOR}x")

    legend = ("R=row  C=col  S=surr  X=diag  ^=up  v=down  <=left  >=right  "
              "↗=NE  ↖=NW  ↘=SE  ↙=SW  "
              "e=evo_greed  o=surr_greed  .=generic_greed")
    print(f"\n  Layout ({legend}):")
    for line in deck.display(display_asgn).splitlines():
        print(f"    {line}")

    if HEATMAP_DISPLAY:
        heatmap = compute_heatmap(deck, asgn, card_class, cores)
        min_r   = min(r for r, _ in deck.slots); max_r = max(r for r, _ in deck.slots)
        min_c   = min(c for _, c in deck.slots); max_c = max(c for _, c in deck.slots)
        cell_w  = max(len(f"{v:.1f}") for v in heatmap.values()) + 1
        print(f"\n  Heatmap (NDM contribution per card):")
        for r in range(min_r, max_r + 1):
            cells = []
            for c in range(min_c, max_c + 1):
                p = (r, c)
                cells.append(f"{heatmap[p]:{cell_w}.1f}" if p in deck.slots else " " * cell_w)
            print(f"    {''.join(cells)}")



# ──────────────────────────────────────────────────────────────────────────────
# Spreadsheet helpers
# ──────────────────────────────────────────────────────────────────────────────

def _lerp_color(
    lo: Tuple[int,int,int], hi: Tuple[int,int,int], t: float,
) -> str:
    r = int(lo[0] + (hi[0]-lo[0]) * t)
    g = int(lo[1] + (hi[1]-lo[1]) * t)
    b = int(lo[2] + (hi[2]-lo[2]) * t)
    return f"{r:02X}{g:02X}{b:02X}"


def _write_balance_block(ws, start_row: int) -> int:
    from openpyxl.styles import Font, PatternFill, Alignment
    HF = PatternFill("solid", fgColor="2C3E50")
    HFo = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    LF = Font(bold=True,  name="Arial", size=10)
    VF = Font(name="Arial", size=10)

    r = start_row
    c = ws.cell(row=r, column=1, value="BALANCE SETTINGS")
    c.font = HFo; c.fill = HF; c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    rows = [
        ("Greed stacking",   "additive" if GREED_ADDITIVE else "multiplicative"),
        ("Core stacking",    "additive" if ADDITIVE_CORES else "multiplicative"),  # ← add this
        ("Exponent active",   "yes" if ENABLE_EXPERIMENTAL_EXPONENT else "no"),
        ("Card exponent",     str(EXPERIMENTAL_EXPONENT) if ENABLE_EXPERIMENTAL_EXPONENT else "n/a"),
        ("Card boost",        str(EXPERIMENTAL_BOOST)    if ENABLE_EXPERIMENTAL_EXPONENT else "n/a"),
        ("foil_core",        f"{MULT_FOIL}x"),
        ("shiny_positional",  "yes" if SHINY_POSITIONAL else "no (T cards, flat 1×)"),
        ("surr_greed",       f"{MULT_SURR_GREED}x"),
        ("dir_greed_vert",   f"{MULT_DIR_GREED_VERT}x"),
        ("dir_greed_horiz",  f"{MULT_DIR_GREED_HORIZ}x"),
        ("dir_greed_diag_up",   f"{MULT_DIR_GREED_DIAG_UP}x   (NE / NW)"),
        ("dir_greed_diag_down", f"{MULT_DIR_GREED_DIAG_DOWN}x   (SE / SW)"),
        ("evo_greed",        f"{MULT_EVO_GREED}x"),
        ("pure_core",        f"{MULT_PURE_BASE} + {MULT_PURE_SCALE} × n_ns"),
        ("equilibrium",      f"{MULT_EQUILIBRIUM}x  (shiny only)"),
        ("steadfast",        f"{MULT_STEADFAST}x  (shiny only)"),
        ("color",            f"{MULT_COLOR}x"),
    ]
    if SPREAD_METRIC:
        rows += [
            ("── Spread metric", ""),
            ("HNS q value",      str(HNS_Q)),
            ("HNS formula",      "N · (Σpᵢ^q)^(1/(1-q));  q=2 → N²/Σhᵢ²"),
        ]
    if ALLOW_DELUXE:
        rows += [
            ("── Deluxe core",    ""),
            ("deluxe_core_base",  f"{MULT_DELUXE_CORE_BASE}"),
            ("deluxe_core_scale", f"+{MULT_DELUXE_CORE_SCALE} per deluxe card"),
            ("deluxe_flat",       f"{MULT_DELUXE_FLAT}x  (flat card multiplier)"),
        ]
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LF
        ws.cell(row=r, column=2, value=value).font = VF
        r += 1
    return r + 1


def _write_overview_block(
    ws,
    all_results: Dict[str, Dict[str, Dict[CardClass, dict]]],
    decks:       List[Deck],
    start_row:   int,
) -> int:
    from openpyxl.styles import Font, PatternFill, Alignment
    HF        = PatternFill("solid", fgColor="2C3E50")
    HFo       = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    SubF      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SubFill_E = PatternFill("solid", fgColor="1E8449")
    SubFill_S = PatternFill("solid", fgColor="1A5276")
    NF        = Font(bold=True, name="Arial", size=10)
    VF        = Font(name="Arial", size=10)

    sample_configs = _get_test_configs(decks[0]) if decks else []
    labels         = [cfg[0] for cfg in sample_configs]
    n_tests        = len(labels)

    evo_start   = 2
    shiny_start = 2 + n_tests
    total_cols  = 1 + 2 * n_tests

    r = start_row

    # Main header
    c = ws.cell(row=r, column=1, value="OVERVIEW — BEST NDM PER DECK")
    c.font = HFo; c.fill = HF; c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    r += 1

    # Class sub-headers
    ws.cell(row=r, column=1, value="").fill = HF
    for label, fill, start_c in [("EVO", SubFill_E, evo_start), ("SHINY", SubFill_S, shiny_start)]:
        end_c = start_c + n_tests - 1
        c = ws.cell(row=r, column=start_c, value=label)
        c.font = SubF; c.fill = fill; c.alignment = Alignment(horizontal="center")
        if n_tests > 1:
            ws.merge_cells(start_row=r, start_column=start_c, end_row=r, end_column=end_c)
    r += 1

    # Test-label row
    ws.cell(row=r, column=1, value="Deck").font = Font(bold=True, name="Arial", size=10)
    for i, lbl in enumerate(labels):
        for offset, fill in [(evo_start, SubFill_E), (shiny_start, SubFill_S)]:
            c = ws.cell(row=r, column=offset + i, value=lbl)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
    r += 1

    # Data rows
    for deck in decks:
        deck_results = all_results.get(deck.name, {})
        ws.cell(row=r, column=1, value=deck.name).font = NF
        for i, (lbl, min_reg, max_greed) in enumerate(_get_test_configs(deck)):
            test_res = deck_results.get(lbl, {})
            cdeck    = deck.with_constraints(min_reg, max_greed)
            for card_class, offset in [(CardClass.EVO, evo_start), (CardClass.SHINY, shiny_start)]:
                if card_class in test_res:
                    ndm = test_res[card_class]["score"]
                    if SPREAD_METRIC:
                        hns = compute_hns_from_result(cdeck, test_res[card_class], card_class)
                        val = f"{ndm:.1f} ({hns:.1f})"
                    else:
                        val = f"{ndm:.1f}"
                else:
                    val = "—"
                c = ws.cell(row=r, column=offset + i, value=val)
                c.font = VF
                c.alignment = Alignment(horizontal="center")
        r += 1

    return r + 1


def _write_class_panel(
    ws,
    deck:       Deck,
    result:     dict,
    card_class: CardClass,
    start_row:  int,
    start_col:  int,
    label:      str = "",
) -> Tuple[int, int]:
    from openpyxl.styles import Font, PatternFill, Alignment

    hdr_color   = "1E8449" if card_class == CardClass.EVO else "1A5276"
    HEADER_FILL = PatternFill("solid", fgColor=hdr_color)
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    SECT_FILL   = PatternFill("solid", fgColor="5D6D7E")
    SECT_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    LABEL_FONT  = Font(bold=True, name="Arial", size=10)
    VALUE_FONT  = Font(name="Arial", size=10)
    GREY_FILL   = PatternFill("solid", fgColor="BDC3C7")
    OUT_FILL    = PatternFill("solid", fgColor="F2F3F4")

    CARD_COLORS = {
        CardType.ROW:             "A9CCE3", CardType.COL:  "A9CCE3", CardType.SURR: "A9CCE3",
        CardType.DIAG: "A9CCE3",
        CardType.DELUXE:          "D7BDE2",   # purple
        CardType.DIR_GREED_UP:    "F9E79F", CardType.DIR_GREED_DOWN:  "F9E79F",
        CardType.DIR_GREED_LEFT:  "F9E79F", CardType.DIR_GREED_RIGHT: "F9E79F",
        CardType.DIR_GREED_NE:    "F9E79F",
        CardType.DIR_GREED_NW:    "F9E79F",
        CardType.DIR_GREED_SE:    "F9E79F",
        CardType.DIR_GREED_SW:    "F9E79F",
        CardType.EVO_GREED:       "FDEBD0", CardType.SURR_GREED:      "FDEBD0",
        CardType.FILLER_GREED:    "D5D8DC",
        CardType.TYPELESS:        "A8D5A2",   # soft green, distinct from blue regular cards
    }

    asgn         = result["assignment"]
    cores        = result["cores"]
    display_asgn = mark_filler_greed(deck, asgn, card_class)
    counts       = Counter(display_asgn.values())
    heatmap      = compute_heatmap(deck, asgn, card_class, cores)

    min_r   = min(r for r, _ in deck.slots); max_r = max(r for r, _ in deck.slots)
    min_c   = min(c for _, c in deck.slots); max_c = max(c for _, c in deck.slots)
    grid_w  = max_c - min_c + 1
    panel_w = max(grid_w, 2)

    def wr(row, col, val=None, *, font=None, fill=None, align=None, nfmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:  c.font          = font
        if fill:  c.fill          = fill
        if align: c.alignment     = align
        if nfmt:  c.number_format = nfmt
        return c

    def merge_header(row, val, fill, font):
        wr(row, start_col, val, font=font, fill=fill, align=Alignment(horizontal="center"))
        if panel_w > 1:
            ws.merge_cells(start_row=row, start_column=start_col,
                           end_row=row, end_column=start_col + panel_w - 1)

    r = start_row

    # Class + test-label header
    header_text = card_class.value.upper()
    if label:
        header_text += f" — {label}"
    merge_header(r, header_text, HEADER_FILL, HEADER_FONT)
    r += 1

    # Info rows
    n_reg     = sum(counts[t] for t in REGULAR_TYPES)
    n_grd     = sum(counts[t] for t in GREED_TYPES)
    breakdown = "  ".join(
        f"{t.value}={counts[t]}" for t in PLACEABLE + [CardType.FILLER_GREED] if counts[t] > 0)

    if SPREAD_METRIC:
        hns       = compute_hns_from_result(deck, result, card_class)
        ndm_label = "NDM  (HNS)"
        ndm_value = f"{result['score']:.4f}  ({hns:.2f})"
    else:
        ndm_label = "NDM"
        ndm_value = f"{result['score']:.4f}"

    for lbl, val in [
        (ndm_label,    ndm_value),
        ("Cores",      ", ".join(sorted(c.value for c in cores))),
        ("Cards",      f"regular={n_reg}  greed={n_grd}"),
        ("Breakdown",  breakdown),
        ("Constraint", deck.constraint_str()),
    ]:
        wr(r, start_col,     lbl, font=LABEL_FONT)
        wr(r, start_col + 1, val, font=VALUE_FONT)
        r += 1

    r += 1  # blank

    # Layout grid
    merge_header(r, "Layout", SECT_FILL, SECT_FONT); r += 1
    for gr in range(min_r, max_r + 1):
        for gc in range(min_c, max_c + 1):
            p       = (gr, gc)
            col_idx = start_col + (gc - min_c)
            if p not in deck.slots:
                wr(r, col_idx, fill=OUT_FILL)
            else:
                ct = display_asgn.get(p, CardType.EMPTY)
                wr(r, col_idx, Deck._CHAR.get(ct, "?"),
                   font=Font(name="Courier New", size=10, bold=ct in REGULAR_TYPES),
                   fill=PatternFill("solid", fgColor=CARD_COLORS.get(ct, "FFFFFF")),
                   align=Alignment(horizontal="center"))
        r += 1

    r += 1  # blank

    # Heatmap grid
    merge_header(r, "Heatmap", SECT_FILL, SECT_FONT); r += 1
    nonzero = [v for v in heatmap.values() if v > 0]
    max_val = max(nonzero) if nonzero else 1.0
    LO = (255, 255, 255); HI = (39, 174, 96)

    for gr in range(min_r, max_r + 1):
        for gc in range(min_c, max_c + 1):
            p       = (gr, gc)
            col_idx = start_col + (gc - min_c)
            if p not in deck.slots:
                wr(r, col_idx, fill=OUT_FILL)
            else:
                val = heatmap[p]
                if val <= 0:
                    wr(r, col_idx, 0, fill=GREY_FILL,
                       font=Font(name="Arial", size=9, color="888888"),
                       align=Alignment(horizontal="center"))
                else:
                    t_val   = val / max_val
                    txt_col = "FFFFFF" if t_val > 0.6 else "000000"
                    wr(r, col_idx, round(val, 1),
                       fill=PatternFill("solid", fgColor=_lerp_color(LO, HI, t_val)),
                       font=Font(name="Arial", size=9, color=txt_col),
                       align=Alignment(horizontal="center"), nfmt="0.0")
        r += 1

    return r - start_row, panel_w


def generate_spreadsheet(
    all_results: Dict[str, Dict[str, Dict[CardClass, dict]]],
    decks:       List[Deck],
    filepath:    str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "NDM Results"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    for i in range(3, 120):
        ws.column_dimensions[get_column_letter(i)].width = 5

    current_row = 1
    current_row = _write_balance_block(ws, current_row)
    current_row = _write_overview_block(ws, all_results, decks, current_row)

    DECK_FILL = PatternFill("solid", fgColor="1C2833")
    DECK_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    GAP       = 2   # blank columns between panels

    for deck in decks:
        deck_results = all_results.get(deck.name, {})
        test_configs = _get_test_configs(deck)

        # Deck header
        ws.cell(row=current_row, column=1,
                value=f"DECK: {deck.name}  ({len(deck.slots)} slots · "
                      f"{deck.core_slots} cores · {deck.n_arcane} arcane)"
                ).font = DECK_FONT
        ws.cell(row=current_row, column=1).fill = DECK_FILL
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=40)
        current_row += 1

        # Compute panel width once (same for all configs since slots are identical)
        min_c  = min(c for _, c in deck.slots); max_c = max(c for _, c in deck.slots)
        grid_w = max_c - min_c + 1
        panel_w = max(grid_w, 2)

        # One row-group per class
        for card_class in CardClass:
            panel_col = 1
            max_rows  = 0

            for label, min_reg, max_greed in test_configs:
                test_result = deck_results.get(label, {})
                if card_class not in test_result:
                    panel_col += panel_w + GAP
                    continue
                cdeck = deck.with_constraints(min_reg, max_greed)
                rows, _ = _write_class_panel(
                    ws, cdeck, test_result[card_class], card_class,
                    current_row, panel_col, label=label)
                max_rows  = max(max_rows, rows)
                panel_col += panel_w + GAP

            current_row += max_rows

        current_row += 2   # blank rows between decks

    wb.save(filepath)
    print(f"\n  Spreadsheet saved → {filepath}")


